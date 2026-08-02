from openpyxl import load_workbook


AI_COLLECTION_CODE = "AI Collection Code"
AI_COLLECTION_NAME = "AI Collection Name"
AI_LANGUAGE = "AI Language"
AI_FORMAT = "AI Format"
AI_AUDIENCE = "AI Audience"
AI_CONFIDENCE = "AI Confidence"
AI_NEEDS_REVIEW = "AI Needs Review"
AI_REASONS = "AI Reasons"

AI_HEADERS = [
    AI_COLLECTION_CODE,
    AI_COLLECTION_NAME,
    AI_LANGUAGE,
    AI_FORMAT,
    AI_AUDIENCE,
    AI_CONFIDENCE,
    AI_NEEDS_REVIEW,
    AI_REASONS,
]


class ExcelWriter:

    def __init__(self, filename):
        self.workbook = load_workbook(filename)
        self.worksheet = self.workbook.active

        self.ai_columns = {}

        self.find_ai_columns()

    def find_ai_columns(self):
        """
        Scan the header row and remember where any AI columns already exist.
        """

        for column in range(1, self.worksheet.max_column + 1):

            header = self.worksheet.cell(row=1, column=column).value

            if header in AI_HEADERS:
                self.ai_columns[header] = column

    def save(self, filename):
        self.workbook.save(filename)

    def write(self, row, column, value):
        self.worksheet.cell(row=row, column=column).value = value

    def write_headers(self):
        """
        Create any missing AI columns.
        If they already exist, leave them alone.
        """

        next_column = self.worksheet.max_column + 1

        for header in AI_HEADERS:

            # Already exists?
            if header in self.ai_columns:
                continue

            # Create it
            self.write(1, next_column, header)

            self.ai_columns[header] = next_column

            next_column += 1

    def write_recommendation(self, row, recommendation):
        self.write(row, self.ai_columns[AI_COLLECTION_CODE], recommendation.collection_code)
        self.write(row, self.ai_columns[AI_COLLECTION_NAME], recommendation.collection_name)
        self.write(row, self.ai_columns[AI_LANGUAGE], recommendation.language)
        self.write(row, self.ai_columns[AI_FORMAT], recommendation.format)
        self.write(row, self.ai_columns[AI_AUDIENCE], recommendation.audience)
        self.write(row, self.ai_columns[AI_CONFIDENCE], recommendation.confidence)
        self.write(row, self.ai_columns[AI_NEEDS_REVIEW], recommendation.needs_review)
        self.write(row, self.ai_columns[AI_REASONS], "; ".join(recommendation.reasons))