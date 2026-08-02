from pathlib import Path
from typing import List

from openpyxl import load_workbook

from cataloger.models.book import Book


def clean_isbn(value) -> str:
    """
    Normalize ISBN values read from Excel.
    """

    if value is None:
        return ""

    value = str(value).strip()

    # Excel sometimes stores text as ="9780064431781"
    if value.startswith('="') and value.endswith('"'):
        value = value[2:-1]

    return value

def load_books(file_path: Path) -> List[Book]:
    """
    Read the workbook and convert every row into a Book object.
    """

    workbook = load_workbook(file_path)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]

    books = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        book = Book(
            book_id=data.get("Book Id"),
            title=data.get("Title") or "",
            author=data.get("Author") or "",
            additional_authors=data.get("Additional Authors") or "",
            isbn=clean_isbn(data.get("ISBN")),
            isbn13=clean_isbn(data.get("ISBN13")),
            publisher=data.get("Publisher") or "",
            year_published=data.get("Year Published"),
            original_publication_year=data.get("Original Publication Year"),
            language=data.get("Language") or "",
            fiction_type=data.get("Fiction/Nonfiction") or "",
            collection=data.get("Collection") or "",
            call_number=data.get("Call Number") or "",
            genre=data.get("Genre") or "",
            review_status=data.get("Review Status") or "Not Started",
        )

        books.append(book)

    return books